# 1. Importaciones de la Librería Estándar
from datetime import timedelta

# 2. Importaciones de Terceros
import openpyxl
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.http import require_POST, require_GET

# 3. Importaciones Locales
from carrera_academica.models import CarreraAcademica
from config.pagination import paginate_queryset
from planta_docente.forms import CargoForm

from planta_docente.models import Asignatura, Cargo, Docente, Resolucion
from planta_docente.utils import (
    calcular_antiguedad,
    calcular_edad,
    formatear_antiguedad,
    obtener_alertas_cargo,
    obtener_estado_jubilacion,
    obtener_estado_vencimiento,
    obtener_cargo_efectivo,
)

# planta_docente/views.py


@login_required
def dashboard_planta_view(request):
    """
    Dashboard principal de planta docente mostrando cargos efectivos.

    Muestra los cargos que los docentes están ejerciendo REALMENTE ahora,
    considerando licencias por mayor jerarquía.

    Args:
        request: HttpRequest

    Query params:
        - q: Búsqueda por nombre/apellido de docente
        - estado: Filtro por estado del cargo
        - departamento: Filtro por departamento
        - dedicacion: Filtro por dedicación
        - categoria: Filtro por categoría
        - alerta: Filtro por tipo de alerta (vencimiento, jubilacion)
        - incluir_inactivos: Incluir cargos de baja, vencidos, jubilados
        - solo_efectivos: Mostrar solo cargos efectivos (excluir bases en lic. M.J.)
        - solo_licencia_mj: Mostrar solo cargos en licencia M.J.
    """

    # Usar manager personalizado
    cargos_qs = Cargo.objects.with_related_data()

    # Aplicar filtros
    search_query = request.GET.get("q", "").strip()
    estado_filter = request.GET.get("estado", "")
    departamento_filter = request.GET.get("departamento", "")
    dedicacion_filter = request.GET.get("dedicacion", "")
    categoria_filter = request.GET.get("categoria", "")
    alerta_filter = request.GET.get("alerta", "")

    # Nuevos filtros para cargos efectivos
    incluir_inactivos = request.GET.get("incluir_inactivos", "false") == "true"

    # Solo efectivos: por defecto TRUE, pero solo si no hay parámetros GET explícitos
    # Si el usuario desmarca, envía el formulario sin el parámetro
    if "solo_efectivos" in request.GET:
        # Usuario interactuó con el checkbox
        solo_efectivos = request.GET.get("solo_efectivos") == "true"
    else:
        # Primera carga o filtro sin interacción → default TRUE
        solo_efectivos = False

    solo_licencia_mj = request.GET.get("solo_licencia_mj", "false") == "true"

    # Filtro de búsqueda
    if search_query:
        cargos_qs = cargos_qs.filter(
            Q(docente__nombre__icontains=search_query)
            | Q(docente__apellido__icontains=search_query)
            | Q(docente__legajo__icontains=search_query)
        )

    # FILTRO PRINCIPAL: Por defecto solo cargos activos efectivos
    if not incluir_inactivos:
        # Excluir cargos de baja
        cargos_qs = cargos_qs.exclude(estado='baja')

        # Excluir docentes jubilados
        cargos_qs = cargos_qs.exclude(docente__jubilado=True)

        # Excluir cargos vencidos
        cargos_qs = cargos_qs.exclude(
            fecha_vencimiento__lt=timezone.now().date()
        )

    # FILTRO CLAVE: Solo efectivos (excluir cargos base en licencia M.J.)
    if solo_efectivos and not solo_licencia_mj:
        # Excluir cargos que están en licencia M.J. (estos son los "base")
        # Solo mostrar los cargos temporales (es_cargo_mayor_jerarquia=True)
        # o los cargos normales (sin licencia M.J.)
        cargos_qs = cargos_qs.exclude(
            en_licencia_mayor_jerarquia=True,
            tipo_cargo_mj='docente'
        )

    # Filtro: Solo licencia M.J. (mostrar SOLO los cargos base en lic. M.J.)
    if solo_licencia_mj:
        cargos_qs = cargos_qs.filter(en_licencia_mayor_jerarquia=True)

    # Filtro por estado (solo si no está activo el filtro de licencia M.J.)
    if estado_filter and not solo_licencia_mj:
        if estado_filter == "activo":
            cargos_qs = cargos_qs.activos()
        elif estado_filter == "licencia":
            cargos_qs = cargos_qs.en_licencia()
        elif estado_filter == "baja":
            cargos_qs = cargos_qs.dados_de_baja()

    # Filtro por departamento
    if departamento_filter:
        cargos_qs = cargos_qs.por_departamento(departamento_filter)

    # Filtro por dedicación
    if dedicacion_filter:
        cargos_qs = cargos_qs.por_dedicacion(dedicacion_filter)

    # Filtro por categoría
    if categoria_filter:
        cargos_qs = cargos_qs.por_categoria(categoria_filter)

    # Filtro por tipo de alerta
    if alerta_filter == "vencimiento":
        cargos_qs = cargos_qs.filter(
            Q(fecha_vencimiento__lte=timezone.now().date() + timedelta(days=180))
            | Q(fecha_vencimiento__lt=timezone.now().date())
        )
    elif alerta_filter == "jubilacion":
        fecha_65_años = timezone.now().date() - timedelta(days=65 * 365)
        cargos_qs = cargos_qs.filter(
            docente__fecha_nacimiento__lte=fecha_65_años)

    # Ordenar por apellido del docente
    cargos_qs = cargos_qs.order_by(
        "docente__apellido", "docente__nombre", "-categoria")

    # ✅ PAGINACIÓN
    page_obj, pagination_context = paginate_queryset(
        cargos_qs, request, page_size=25)

    # Enriquecer cada cargo con información calculada
    for cargo in page_obj:
        # ✅ NUEVO: Obtener cargo efectivo
        cargo.cargo_efectivo = obtener_cargo_efectivo(cargo)

        # Estado de vencimiento
        cargo.estado_venc = obtener_estado_vencimiento(cargo)

        # Estado de jubilación del docente
        cargo.estado_jub = obtener_estado_jubilacion(cargo.docente)

        # Antigüedad
        antiguedad = calcular_antiguedad(cargo.fecha_inicio)
        cargo.antiguedad_texto = formatear_antiguedad(antiguedad)
        cargo.antiguedad_años = antiguedad["años"]

        # Alertas
        cargo.alertas = obtener_alertas_cargo(cargo)
        cargo.tiene_alertas_urgentes = any(a["urgente"] for a in cargo.alertas)

        # Verificar si tiene CA
        cargo.tiene_ca = hasattr(cargo, "carrera_academica")

    # Calcular estadísticas generales (solo para cargos activos efectivos)
    stats = _calcular_estadisticas_dashboard(
        cargos_qs, incluir_inactivos, solo_efectivos)

    # Preparar contexto
    contexto = {
        "cargos": page_obj,
        "stats": stats,
        "search_query": search_query,
        "estado_filter": estado_filter,
        "departamento_filter": departamento_filter,
        "dedicacion_filter": dedicacion_filter,
        "categoria_filter": categoria_filter,
        "alerta_filter": alerta_filter,
        "incluir_inactivos": incluir_inactivos,
        "solo_efectivos": solo_efectivos,
        "solo_licencia_mj": solo_licencia_mj,
        "estado_choices": Cargo.ESTADO_CHOICES,
        "departamento_choices": Asignatura.DEPTO_CHOICES,
        "dedicacion_choices": Cargo.DEDICACION_CHOICES,
        "categoria_choices": Cargo.CATEGORIA_CHOICES,
        **pagination_context,
    }

    return render(request, "planta_docente/dashboard.html", contexto)


def _calcular_estadisticas_dashboard(cargos_qs, incluir_inactivos=False, solo_efectivos=True):
    """
    Calcula estadísticas para el dashboard basadas en cargos efectivos.

    Args:
        cargos_qs: QuerySet de cargos (ya filtrado)
        incluir_inactivos: Si se están incluyendo cargos inactivos
        solo_efectivos: Si solo se muestran cargos efectivos

    Returns:
        dict: Diccionario con estadísticas
    """
    # Base queryset para stats (siempre excluir inactivos en stats)
    stats_qs = cargos_qs.exclude(estado='baja').exclude(
        docente__jubilado=True
    ).exclude(
        fecha_vencimiento__lt=timezone.now().date()
    )

    # Si solo_efectivos, excluir cargos base en lic. M.J.
    if solo_efectivos:
        stats_qs = stats_qs.exclude(
            en_licencia_mayor_jerarquia=True,
            tipo_cargo_mj='docente'
        )

    # Total de cargos efectivos
    total_cargos = stats_qs.count()

    # Cargos activos (sin considerar licencias)
    cargos_activos = stats_qs.filter(estado="activo").count()

    # Cargos en licencia
    cargos_licencia = stats_qs.filter(estado="licencia").count()

    # Cargos en licencia M.J. (para la stat específica, contar los base)
    cargos_licencia_mj = Cargo.objects.filter(
        en_licencia_mayor_jerarquia=True
    ).exclude(
        docente__jubilado=True
    ).exclude(
        fecha_vencimiento__lt=timezone.now().date()
    ).count()

    # Cargos próximos a vencer (60 días) - solo activos efectivos
    cargos_criticos = stats_qs.filter(
        fecha_vencimiento__lte=timezone.now().date() + timedelta(days=60),
        fecha_vencimiento__gte=timezone.now().date(),
        estado='activo'
    ).count()

    # Cargos vencidos (solo si se incluyen inactivos)
    if incluir_inactivos:
        cargos_vencidos = cargos_qs.filter(
            fecha_vencimiento__lt=timezone.now().date()
        ).count()

        total_docentes_jubilados = (
            cargos_qs.filter(docente__jubilado=True).values(
                "docente").distinct().count()
        )
    else:
        cargos_vencidos = 0
        total_docentes_jubilados = 0

    # Cargos regulares/ordinarios sin CA (siempre sobre activos efectivos)
    cargos_sin_ca = stats_qs.filter(
        caracter__in=["reg", "ord"],
        carrera_academica__isnull=True,
        estado='activo'
    ).count()

    # Docentes únicos con alerta de jubilación (siempre sobre activos efectivos)
    fecha_65_años = timezone.now().date() - timedelta(days=65 * 365)
    docentes_mayores_65 = (
        stats_qs.filter(
            docente__fecha_nacimiento__lte=fecha_65_años,
            estado='activo'
        )
        .values("docente")
        .distinct()
        .count()
    )

    return {
        "total_cargos": total_cargos,
        "cargos_activos": cargos_activos,
        "cargos_licencia": cargos_licencia,
        "cargos_licencia_mj": cargos_licencia_mj,
        "cargos_criticos": cargos_criticos,
        "cargos_vencidos": cargos_vencidos,
        "cargos_sin_ca": cargos_sin_ca,
        "docentes_mayores_65": docentes_mayores_65,
        "total_docentes_jubilados": total_docentes_jubilados,
    }


@login_required
def detalle_cargo_view(request, pk):
    """
    Vista de detalle de un cargo específico.

    Muestra:
    - Información completa del cargo
    - Información del docente
    - Antigüedad calculada
    - Estado de vencimiento
    - Estado de jubilación del docente
    - Historial de resoluciones
    - Carrera Académica asociada (si existe)
    - Botón para iniciar CA (si corresponde)

    Args:
        request: HttpRequest
        pk: ID del cargo
    """
    # ✅ OPTIMIZACIÓN: Usar select_related y prefetch_related
    cargo = get_object_or_404(
        Cargo.objects.select_related(
            "docente",
            "asignatura",
        ).prefetch_related(
            "docente__correos",
            "resoluciones",
        ),
        pk=pk,
    )

    # Calcular información adicional
    estado_venc = obtener_estado_vencimiento(cargo)
    estado_jub = obtener_estado_jubilacion(cargo.docente)
    antiguedad = calcular_antiguedad(cargo.fecha_inicio)
    antiguedad_texto = formatear_antiguedad(antiguedad)
    alertas = obtener_alertas_cargo(cargo)
    
    #Licencias
    estado_licencia = cargo.get_estado_licencia_display()
    info_continuidad = cargo.get_info_continuidad()

    # Verificar si tiene CA
    tiene_ca = hasattr(cargo, "carrera_academica")
    puede_iniciar_ca = (
        cargo.caracter in ["reg", "ord"] 
        and not tiene_ca 
        and cargo.estado != 'baja'
        and not cargo.docente.jubilado
    )

    # Obtener correo principal del docente
    correo_principal = cargo.docente.correos.filter(principal=True).first()

    # Historial de resoluciones ordenado
    resoluciones = cargo.resoluciones.all().order_by("-año", "-numero")

    contexto = {
        "cargo": cargo,
        "estado_venc": estado_venc,
        "estado_jub": estado_jub,
        "antiguedad": antiguedad,
        "antiguedad_texto": antiguedad_texto,
        "alertas": alertas,
        "tiene_ca": tiene_ca,
        "puede_iniciar_ca": puede_iniciar_ca,
        "correo_principal": correo_principal,
        'estado_licencia': estado_licencia,
        'info_continuidad': info_continuidad,
        "resoluciones": resoluciones,
    }

    return render(request, "planta_docente/cargo_detail.html", contexto)


@login_required
def detalle_docente_view(request, pk):
    """
    Vista de detalle de un docente.

    Muestra:
    - Información personal
    - Edad y estado de jubilación
    - Lista de todos sus cargos (activos e históricos)
    - Correos electrónicos

    Args:
        request: HttpRequest
        pk: ID del docente
    """
    # ✅ OPTIMIZACIÓN: Prefetch de relaciones
    docente = get_object_or_404(
        Docente.objects.prefetch_related(
            "correos",
            "cargo_docente",
            "cargo_docente__asignatura",
            "cargo_docente__carrera_academica",
        ),
        pk=pk,
    )

    # Calcular edad y estado de jubilación
    edad = calcular_edad(docente.fecha_nacimiento)
    estado_jub = obtener_estado_jubilacion(docente)

    # Separar cargos activos e históricos
    cargos_activos = docente.cargo_docente.filter(estado="activo")
    cargos_historicos = docente.cargo_docente.exclude(estado="activo")

    # Enriquecer cargos activos con información
    for cargo in cargos_activos:
        cargo.estado_venc = obtener_estado_vencimiento(cargo)
        antiguedad = calcular_antiguedad(cargo.fecha_inicio)
        cargo.antiguedad_texto = formatear_antiguedad(antiguedad)
        cargo.tiene_ca = hasattr(cargo, "carrera_academica")

    # Correos
    correo_principal = docente.correos.filter(principal=True).first()
    otros_correos = docente.correos.filter(principal=False)

    contexto = {
        "docente": docente,
        "edad": edad,
        "estado_jub": estado_jub,
        "cargos_activos": cargos_activos,
        "cargos_historicos": cargos_historicos,
        "correo_principal": correo_principal,
        "otros_correos": otros_correos,
    }

    return render(request, "planta_docente/docente_detail.html", contexto)


@login_required
def vencimientos_view(request):
    """
    Reporte de cargos próximos a vencer.

    Muestra cargos que vencen en los próximos 6 meses,
    ordenados por fecha de vencimiento.

    Query params:
        - dias: Días hacia adelante (default: 180)
        - incluir_vencidos: Si incluir cargos ya vencidos (default: True)
    """
    dias = int(request.GET.get("dias", 180))
    incluir_vencidos = request.GET.get("incluir_vencidos", "true") == "true"

    # Obtener cargos próximos a vencer
    cargos_qs = Cargo.objects.with_related_data().proximos_a_vencer(dias)

    # Incluir vencidos si se solicita
    if incluir_vencidos:
        cargos_vencidos = Cargo.objects.with_related_data().vencidos()
        cargos_qs = cargos_qs | cargos_vencidos

    # Ordenar por urgencia (vencidos primero, luego por fecha)
    cargos_qs = cargos_qs.order_by("fecha_vencimiento")

    # Enriquecer con información
    cargos = list(cargos_qs)
    for cargo in cargos:
        cargo.estado_venc = obtener_estado_vencimiento(cargo)
        antiguedad = calcular_antiguedad(cargo.fecha_inicio)
        cargo.antiguedad_texto = formatear_antiguedad(antiguedad)
        cargo.tiene_ca = hasattr(cargo, "carrera_academica")

    # Agrupar por urgencia
    vencidos = [c for c in cargos if c.estado_venc["tipo"] == "vencido"]
    criticos = [c for c in cargos if c.estado_venc["tipo"] == "critico"]
    proximos = [c for c in cargos if c.estado_venc["tipo"] == "proximo"]

    contexto = {
        "vencidos": vencidos,
        "criticos": criticos,
        "proximos": proximos,
        "total_alertas": len(vencidos) + len(criticos) + len(proximos),
        "dias": dias,
        "incluir_vencidos": incluir_vencidos,
    }

    return render(request, "planta_docente/vencimientos.html", contexto)


@login_required
def jubilaciones_view(request):
    """
    Reporte de docentes próximos a jubilarse.

    Muestra docentes que cumplen 65 o 70 años en los próximos años.

    Query params:
        - años: Años hacia adelante (default: 2)
    """
    años = int(request.GET.get("años", 2))

    # Obtener docentes próximos a jubilarse
    docentes_qs = Docente.objects.with_related_data().proximos_a_jubilarse(años)

    # También incluir los que ya tienen 65+
    docentes_mayores = Docente.objects.with_related_data().mayores_de_65()

    # Unir querysets
    docentes_qs = (docentes_qs | docentes_mayores).distinct()

    # Enriquecer con información
    docentes = list(docentes_qs)
    for docente in docentes:
        docente.edad = calcular_edad(docente.fecha_nacimiento)
        docente.estado_jub = obtener_estado_jubilacion(docente)

        # Obtener cargos activos
        docente.cargos_activos = docente.cargo_docente.filter(estado="activo")
        for cargo in docente.cargos_activos:
            cargo.tiene_ca = hasattr(cargo, "carrera_academica")

    # Agrupar por urgencia
    mayores_70 = [d for d in docentes if d.estado_jub["estado"] == "jubilado_70"]
    entre_65_70 = [d for d in docentes if d.estado_jub["estado"] == "jubilado_65"]
    proximos_65 = [d for d in docentes if d.estado_jub["estado"] == "proximo_65"]

    contexto = {
        "mayores_70": mayores_70,
        "entre_65_70": entre_65_70,
        "proximos_65": proximos_65,
        "total_alertas": len(mayores_70) + len(entre_65_70) + len(proximos_65),
        "años": años,
    }

    return render(request, "planta_docente/jubilaciones.html", contexto)


@login_required
def cargos_sin_ca_view(request):
    """
    Reporte de cargos regulares/ordinarios sin Carrera Académica.

    Lista cargos que deberían tener CA iniciada pero no la tienen.
    """
    # Obtener cargos sin CA
    cargos_qs = Cargo.objects.with_related_data().sin_ca().filter(estado="activo")

    # Ordenar por antigüedad (más antiguos primero)
    cargos_qs = cargos_qs.order_by("fecha_inicio")

    # Enriquecer con información
    cargos = list(cargos_qs)
    for cargo in cargos:
        antiguedad = calcular_antiguedad(cargo.fecha_inicio)
        cargo.antiguedad_texto = formatear_antiguedad(antiguedad)
        cargo.antiguedad_años = antiguedad["años"]
        cargo.estado_venc = obtener_estado_vencimiento(cargo)

    contexto = {
        "cargos": cargos,
        "total_sin_ca": len(cargos),
    }

    return render(request, "planta_docente/sin_ca.html", contexto)


@login_required
def iniciar_ca_desde_cargo_view(request, pk):
    """
    Inicia una Carrera Académica desde un cargo existente.

    Redirige al formulario de creación de CA con el cargo preseleccionado.

    Args:
        request: HttpRequest
        pk: ID del cargo
    """
    cargo = get_object_or_404(Cargo, pk=pk)

    # Verificar que el cargo puede tener CA
    if cargo.caracter not in ["reg", "ord"]:
        messages.error(
            request,
            f"Solo los cargos Regulares u Ordinarios pueden tener Carrera Académica. "
            f"Este cargo es {cargo.get_caracter_display()}.",
        )
        return redirect("planta_docente:detalle_cargo", pk=pk)

    # Verificar que no tenga CA ya
    if hasattr(cargo, "carrera_academica"):
        messages.warning(
            request, "Este cargo ya tiene una Carrera Académica iniciada.")
        return redirect("carrera_academica:detalle_ca", pk=cargo.carrera_academica.pk)

    # Solo verificar que NO esté de baja y que el docente NO esté jubilado
    if cargo.estado == 'baja':
        messages.error(
            request,
            "No se puede iniciar Carrera Académica para un cargo dado de baja.",
        )
        return redirect("planta_docente:detalle_cargo", pk=pk)

    if cargo.docente.jubilado:
        messages.error(
            request,
            f"No se puede iniciar Carrera Académica para un docente jubilado ({cargo.docente}).",
        )
        return redirect("planta_docente:detalle_cargo", pk=pk)

    if request.method == "POST":
        # Crear la CA con las fechas del cargo
        try:
            nueva_ca = CarreraAcademica(
                cargo=cargo,
                fecha_inicio=cargo.fecha_inicio,
                fecha_vencimiento_original=cargo.fecha_vencimiento,
                fecha_vencimiento_actual=cargo.fecha_vencimiento,
            )
            nueva_ca.full_clean()
            nueva_ca.save()

            messages.success(
                request,
                f"Carrera Académica iniciada exitosamente para {cargo.docente}.",
            )
            return redirect("carrera_academica:detalle_ca", pk=nueva_ca.pk)

        except Exception as e:
            messages.error(request, f"Error al crear la Carrera Académica: {str(e)}")
            return redirect("planta_docente:detalle_cargo", pk=pk)

    # GET: Mostrar confirmación
    contexto = {
        "cargo": cargo,
        "docente": cargo.docente,
        "asignatura": cargo.asignatura,
    }

    return render(request, "planta_docente/confirmar_iniciar_ca.html", contexto)


@login_required
def cargo_info_api_view(request, pk):
    """
    API endpoint para obtener información de un cargo en formato JSON.

    Útil para cargar datos vía AJAX sin recargar la página.

    Args:
        request: HttpRequest
        pk: ID del cargo

    Returns:
        JsonResponse con información del cargo
    """
    try:
        cargo = Cargo.objects.select_related("docente", "asignatura").get(pk=pk)

        # Calcular información
        estado_venc = obtener_estado_vencimiento(cargo)
        estado_jub = obtener_estado_jubilacion(cargo.docente)
        antiguedad = calcular_antiguedad(cargo.fecha_inicio)

        data = {
            "id": cargo.pk,
            "docente": {
                "nombre_completo": str(cargo.docente),
                "legajo": cargo.docente.legajo,
                "edad": calcular_edad(cargo.docente.fecha_nacimiento),
            },
            "asignatura": {
                "nombre": cargo.asignatura.nombre,
                "departamento": cargo.get_asignatura.get_departamento_display(),
            },
            "cargo": {
                "caracter": cargo.get_caracter_display(),
                "categoria": cargo.get_categoria_display(),
                "dedicacion": cargo.get_dedicacion_display(),
                "estado": cargo.get_estado_display(),
            },
            "fechas": {
                "inicio": cargo.fecha_inicio.strftime("%d/%m/%Y"),
                "vencimiento": (
                    cargo.fecha_vencimiento.strftime("%d/%m/%Y")
                    if cargo.fecha_vencimiento
                    else None
                ),
            },
            "antiguedad": {
                "años": antiguedad["años"],
                "meses": antiguedad["meses"],
                "dias": antiguedad["dias"],
                "texto": formatear_antiguedad(antiguedad),
            },
            "estado_vencimiento": estado_venc,
            "estado_jubilacion": estado_jub,
            "tiene_ca": hasattr(cargo, "carrera_academica"),
        }

        return JsonResponse(data)

    except Cargo.DoesNotExist:
        return JsonResponse({"error": "Cargo no encontrado"}, status=404)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@login_required
@staff_member_required
def cargos_renovacion_view(request):
    """Vista para gestionar renovaciones de cargos interinos y ad-honorem."""

    # Obtener cargos interinos y ad-honorem ACTIVOS
    cargos_qs = (
        Cargo.objects.filter(
            caracter__in=["int", "adh"],
            estado="activo",
            docente__jubilado=False,  # Excluir docentes jubilados
        )
        .select_related(
            "docente", "asignatura", "carrera_academica", "usuario_renovacion"
        )
        .order_by(
            "renovacion_solicitada",  # Primero los no renovados
            "docente__apellido",
            "docente__nombre",
        )
    )

    # Agrupar por docente
    from collections import defaultdict

    docentes_dict = defaultdict(list)

    total_renovados = 0
    total_pendientes = 0

    for cargo in cargos_qs:
        docentes_dict[cargo.docente].append(cargo)
        if cargo.renovacion_solicitada:
            total_renovados += 1
        else:
            total_pendientes += 1

    contexto = {
        "docentes_con_cargos": dict(docentes_dict),
        "total_cargos": cargos_qs.count(),
        "total_docentes": len(docentes_dict),
        "total_renovados": total_renovados,
        "total_pendientes": total_pendientes,
    }

    return render(request, "planta_docente/cargos_renovacion.html", contexto)


@login_required
@staff_member_required
@require_POST
def renovar_cargo_ajax(request, cargo_id):
    """Vista AJAX para renovar/cancelar renovación de un cargo."""

    cargo = get_object_or_404(Cargo, pk=cargo_id)
    accion = request.POST.get("accion", "renovar")

    if accion == "renovar":
        exito, mensaje = cargo.solicitar_renovacion(request.user)
    elif accion == "cancelar":
        exito, mensaje = cargo.cancelar_renovacion()
    else:
        return JsonResponse(
            {"success": False, "message": "Acción no válida"}, status=400
        )

    return JsonResponse(
        {
            "success": exito,
            "message": mensaje,
            "renovado": cargo.renovacion_solicitada,
            "fecha_vencimiento": (
                cargo.fecha_vencimiento.strftime("%d/%m/%Y")
                if cargo.fecha_vencimiento
                else None
            ),
            "fecha_renovacion": (
                cargo.fecha_renovacion.strftime("%d/%m/%Y")
                if cargo.fecha_renovacion
                else None
            ),
            "usuario": (
                str(cargo.usuario_renovacion) if cargo.usuario_renovacion else None
            ),
        }
    )


@login_required
@staff_member_required
def exportar_renovaciones_excel(request):
    """Exporta todos los cargos a renovar a Excel."""

    # Obtener cargos
    cargos_qs = (
        Cargo.objects.filter(
            caracter__in=["int", "adh"], estado="activo", docente__jubilado=False
        )
        .select_related("docente", "asignatura")
        .order_by("renovacion_solicitada", "docente__apellido", "docente__nombre")
    )

    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cargos a Renovar"

    # Headers
    headers = [
        "Apellido y Nombre",
        "Legajo",
        "Cargo y Jerarquía",
        "Dedicación",
        "Carácter",
        "Asignatura",
        "Renovación Solicitada",
    ]

    # Escribir headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)

    # Datos
    for row_num, cargo in enumerate(cargos_qs, 2):
        # Apellido y Nombre
        ws.cell(
            row=row_num,
            column=1,
            value=f"{cargo.docente.apellido}, {cargo.docente.nombre}",
        )

        # Legajo
        ws.cell(row=row_num, column=2, value=cargo.docente.legajo)

        # Cargo y Jerarquía
        ws.cell(row=row_num, column=3, value=cargo.get_jerarquia_display())

        # Dedicación
        ws.cell(row=row_num, column=4, value=cargo.get_dedicacion_display())

        # Carácter
        ws.cell(row=row_num, column=5, value=cargo.get_caracter_display())

        # Asignatura
        ws.cell(
            row=row_num,
            column=6,
            value=cargo.asignatura.nombre if cargo.asignatura else "-",
        )

        # Renovación Solicitada
        ws.cell(
            row=row_num, column=7, value="Sí" if cargo.renovacion_solicitada else "No"
        )

    # Ajustar anchos
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 2

    # Preparar response
    from django.utils import timezone

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f'cargos_renovacion_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


@login_required
@staff_member_required
def crear_cargo_view(request):
    """Vista para crear un nuevo cargo docente."""

    if request.method == 'POST':
        form = CargoForm(request.POST)

        if form.is_valid():
            cargo = form.save()

            messages.success(
                request,
                f'✓ Cargo creado exitosamente: {cargo.docente.apellido}, {cargo.docente.nombre} - '
                f'{cargo.get_categoria_display()} {cargo.get_caracter_display()}'
            )

            return redirect('planta_docente:detalle_cargo', pk=cargo.pk)
        else:
            messages.error(
                request, 'Por favor corrige los errores del formulario.')
    else:
        form = CargoForm()

    contexto = {
        'form': form,
        'title': 'Nuevo Cargo Docente',
        'is_new': True
    }

    return render(request, 'planta_docente/cargo_form.html', contexto)


@login_required
@staff_member_required
def editar_cargo_view(request, pk):
    """Vista para editar un cargo existente."""

    cargo = get_object_or_404(Cargo, pk=pk)

    if request.method == 'POST':
        form = CargoForm(request.POST, instance=cargo)

        if form.is_valid():
            cargo = form.save()

            messages.success(
                request,
                f'✓ Cargo actualizado exitosamente: {cargo.docente.apellido}, {cargo.docente.nombre} - '
                f'{cargo.get_categoria_display()}'
            )

            return redirect('planta_docente:detalle_cargo', pk=cargo.pk)
        else:
            messages.error(
                request, 'Por favor corrige los errores del formulario.')
    else:
        form = CargoForm(instance=cargo)

    contexto = {
        'form': form,
        'cargo': cargo,
        'title': 'Editar Cargo',
        'is_new': False
    }

    return render(request, 'planta_docente/cargo_form.html', contexto)


@login_required
@staff_member_required
def gestionar_resoluciones_cargo(request, pk):
    """Vista para gestionar resoluciones de un cargo."""
    cargo = get_object_or_404(Cargo, pk=pk)

    # Resoluciones de este cargo
    resoluciones_cargo = cargo.resoluciones.all().order_by('-año', '-numero')

    # Todas las resoluciones para poder crear nuevas
    # (esto se usará en el formulario de crear)

    contexto = {
        'cargo': cargo,
        'resoluciones': resoluciones_cargo,
    }

    return render(request, 'planta_docente/gestionar_resoluciones.html', contexto)


@login_required
@staff_member_required
def crear_resolucion_cargo(request, pk):
    """Vista para crear una nueva resolución para un cargo."""
    cargo = get_object_or_404(Cargo, pk=pk)

    if request.method == 'POST':
        # Obtener datos del formulario
        numero = request.POST.get('numero')
        año = request.POST.get('año')
        objeto = request.POST.get('objeto')
        origen = request.POST.get('origen')
        file = request.FILES.get('file')

        # Crear resolución
        try:
            resolucion = Resolucion.objects.create(
                cargo=cargo,
                numero=int(numero),
                año=int(año),
                objeto=objeto,
                origen=origen,
                file=file if file else None
            )

            messages.success(
                request,
                f'✓ Resolución {resolucion.get_origen_display()} {numero}/{año} creada exitosamente.'
            )

            return redirect('planta_docente:gestionar_resoluciones_cargo', pk=pk)

        except Exception as e:
            messages.error(request, f'Error al crear resolución: {str(e)}')

    contexto = {
        'cargo': cargo,
        'objeto_choices': Resolucion.OBJETO_CHOICES,
        'origen_choices': Resolucion.ORIGEN_CHOICES,
    }

    return render(request, 'planta_docente/resolucion_create.html', contexto)


@login_required
@staff_member_required
@require_POST
def eliminar_resolucion_cargo(request, cargo_pk, resolucion_pk):
    """Eliminar una resolución del cargo."""
    cargo = get_object_or_404(Cargo, pk=cargo_pk)
    resolucion = get_object_or_404(Resolucion, pk=resolucion_pk, cargo=cargo)

    # Guardar info antes de eliminar
    info_res = f'{resolucion.get_origen_display()} {resolucion.numero}/{resolucion.año}'

    # Eliminar
    resolucion.delete()

    messages.warning(
        request,
        f'Resolución {info_res} eliminada.'
    )

    return redirect('planta_docente:gestionar_resoluciones_cargo', pk=cargo_pk)


@login_required
@staff_member_required
def gestionar_licencia_cargo(request, pk):
    """Vista para gestionar licencias (normal y mayor jerarquía)."""
    cargo = get_object_or_404(Cargo, pk=pk)

    if request.method == 'POST':
        tipo_licencia = request.POST.get('tipo_licencia')
        accion = request.POST.get('accion')

        try:
            from datetime import datetime

            # ====================================
            # LICENCIA NORMAL
            # ====================================
            if tipo_licencia == 'normal':
                if accion == 'alta':
                    fecha_inicio = datetime.strptime(
                        request.POST.get('fecha_inicio'), '%Y-%m-%d').date()
                    fecha_fin = datetime.strptime(
                        request.POST.get('fecha_fin'), '%Y-%m-%d').date()

                    exito, mensaje = cargo.dar_alta_licencia_normal(
                        fecha_inicio=fecha_inicio,
                        fecha_fin=fecha_fin,
                        usuario=request.user
                    )

                    if exito:
                        messages.success(request, mensaje)

                        # Crear resolución si se proporcionan datos
                        numero = request.POST.get('numero_resolucion')
                        año = request.POST.get('año_resolucion')
                        origen = request.POST.get('origen_resolucion')

                        if numero and año and origen:
                            Resolucion.objects.create(
                                cargo=cargo,
                                numero=int(numero),
                                año=int(año),
                                objeto='licencia_alta',
                                origen=origen,
                                fecha_inicio_licencia=fecha_inicio,
                                fecha_fin_licencia=fecha_fin,
                                genera_prorroga_ca=False  # Licencia normal NO genera prórroga
                            )
                    else:
                        messages.error(request, mensaje)

                elif accion == 'baja':
                    exito, mensaje = cargo.dar_baja_licencia_normal()

                    if exito:
                        messages.success(request, mensaje)
                    else:
                        messages.error(request, mensaje)

            # ====================================
            # LICENCIA POR MAYOR JERARQUÍA
            # ====================================
            elif tipo_licencia == 'mayor_jerarquia':
                if accion == 'alta':
                    fecha_inicio = datetime.strptime(
                        request.POST.get('fecha_inicio'), '%Y-%m-%d').date()

                    exito, mensaje = cargo.dar_alta_licencia_mayor_jerarquia(
                        fecha_inicio=fecha_inicio,
                        usuario=request.user
                    )

                    if exito:
                        messages.success(request, mensaje)

                        # Crear resolución si se proporcionan datos
                        numero = request.POST.get('numero_resolucion')
                        año = request.POST.get('año_resolucion')
                        origen = request.POST.get('origen_resolucion')

                        if numero and año and origen:
                            Resolucion.objects.create(
                                cargo=cargo,
                                numero=int(numero),
                                año=int(año),
                                objeto='licencia_alta',
                                origen=origen,
                                fecha_inicio_licencia=fecha_inicio,
                                genera_prorroga_ca=True  # Licencia M.J. SÍ genera prórroga
                            )
                    else:
                        messages.error(request, mensaje)

                elif accion == 'baja':
                    fecha_fin = datetime.strptime(
                        request.POST.get('fecha_fin'), '%Y-%m-%d').date()

                    exito, mensaje = cargo.dar_baja_licencia_mayor_jerarquia(
                        fecha_fin=fecha_fin,
                        usuario=request.user
                    )

                    if exito:
                        messages.success(request, mensaje)

                        # Crear resolución si se proporcionan datos
                        numero = request.POST.get('numero_resolucion')
                        año = request.POST.get('año_resolucion')
                        origen = request.POST.get('origen_resolucion')

                        if numero and año and origen:
                            Resolucion.objects.create(
                                cargo=cargo,
                                numero=int(numero),
                                año=int(año),
                                objeto='licencia_baja',
                                origen=origen,
                                fecha_fin_licencia=fecha_fin,
                                genera_prorroga_ca=True
                            )
                    else:
                        messages.error(request, mensaje)

        except Exception as e:
            messages.error(request, f'Error al procesar licencia: {str(e)}')

        return redirect('planta_docente:detalle_cargo', pk=pk)

    # GET request
    estado_licencia = cargo.get_estado_licencia_display()

    contexto = {
        'cargo': cargo,
        'estado_licencia': estado_licencia,
        'origen_choices': Resolucion.ORIGEN_CHOICES,
    }

    return render(request, 'planta_docente/gestionar_licencia.html', contexto)


@login_required
@staff_member_required
def gestionar_continuidad_cargo(request, pk):
    """Vista para gestionar la continuidad de un cargo."""
    cargo = get_object_or_404(Cargo, pk=pk)

    if request.method == 'POST':
        accion = request.POST.get('accion')

        try:
            if accion == 'finalizar_sin_continuidad':
                razon = request.POST.get('razon_finalizacion')
                observaciones = request.POST.get('observaciones')

                exito, mensaje = cargo.finalizar_sin_continuidad(
                    razon=razon,
                    observaciones=observaciones,
                    usuario=request.user
                )

                if exito:
                    messages.success(request, mensaje)
                else:
                    messages.error(request, mensaje)

            elif accion == 'finalizar_con_continuidad':
                opcion_cargo = request.POST.get('opcion_cargo')

                # ✅ OPCIÓN 1: Cargo existente
                if opcion_cargo == 'existente':
                    cargo_siguiente_id = request.POST.get('cargo_siguiente_id')

                    if not cargo_siguiente_id:
                        messages.error(
                            request, 'Debe seleccionar un cargo siguiente')
                        return redirect('planta_docente:gestionar_continuidad_cargo', pk=pk)

                    cargo_siguiente = get_object_or_404(
                        Cargo, pk=cargo_siguiente_id)

                # ✅ OPCIÓN 2: Crear nuevo cargo
                elif opcion_cargo == 'nuevo':
                    # Obtener datos del nuevo cargo
                    categoria = request.POST.get('nueva_categoria')
                    dedicacion = request.POST.get('nueva_dedicacion')
                    tipo_cargo = request.POST.get('nuevo_tipo_cargo')
                    fecha_designacion_str = request.POST.get(
                        'nueva_fecha_designacion')

                    # Validar campos
                    if not all([categoria, dedicacion, tipo_cargo, fecha_designacion_str]):
                        messages.error(
                            request, 'Todos los campos del nuevo cargo son obligatorios')
                        return redirect('planta_docente:gestionar_continuidad_cargo', pk=pk)

                    # Parsear fecha
                    from datetime import datetime
                    fecha_designacion = datetime.strptime(
                        fecha_designacion_str, '%Y-%m-%d').date()

                    # Crear el nuevo cargo
                    cargo_siguiente = Cargo.objects.create(
                        docente=cargo.docente,
                        asignatura=cargo.asignatura,  # ✅ Hereda asignatura
                        categoria=categoria,
                        dedicacion=dedicacion,
                        tipo_cargo=tipo_cargo,
                        caracter=tipo_cargo,  # Usamos tipo_cargo como caracter
                        fecha_inicio=fecha_designacion,
                        fecha_designacion=fecha_designacion,
                        estado='activo',
                        estado_continuidad='activo'
                    )

                    messages.info(
                        request, f'✓ Nuevo cargo creado: {cargo_siguiente.get_categoria_display()}')
                else:
                    messages.error(request, 'Opción de cargo no válida')
                    return redirect('planta_docente:gestionar_continuidad_cargo', pk=pk)

                # Continuar con la vinculación
                tipo_continuidad = request.POST.get('tipo_continuidad')
                observaciones = request.POST.get('observaciones')

                exito, mensaje = cargo.finalizar_con_continuidad(
                    cargo_siguiente=cargo_siguiente,
                    tipo_continuidad=tipo_continuidad,
                    observaciones=observaciones,
                    usuario=request.user
                )

                if exito:
                    messages.success(request, mensaje)
                else:
                    messages.error(request, mensaje)

            elif accion == 'desvincular':
                exito, mensaje = cargo.desvincular_continuidad()

                if exito:
                    messages.success(request, mensaje)
                else:
                    messages.error(request, mensaje)

        except Exception as e:
            messages.error(request, f'Error al procesar continuidad: {str(e)}')

        return redirect('planta_docente:detalle_cargo', pk=pk)

    # GET request
    info_continuidad = cargo.get_info_continuidad()
    cadena = cargo.obtener_cadena_continuidad()

    # Obtener cargos posibles para continuidad (mismo docente, activos)
    cargos_posibles = Cargo.objects.filter(
        docente=cargo.docente,
        estado='activo',
        estado_continuidad='activo'
    ).exclude(pk=cargo.pk).order_by('-fecha_inicio')

    contexto = {
        'cargo': cargo,
        'info_continuidad': info_continuidad,
        'cadena': cadena,
        'cargos_posibles': cargos_posibles,
        'razon_choices': Cargo.RAZON_FINALIZACION_CHOICES,
        'tipo_continuidad_choices': Cargo.TIPO_CONTINUIDAD_CHOICES,
        'categoria_choices': Cargo.CATEGORIA_CHOICES,
        'dedicacion_choices': Cargo.DEDICACION_CHOICES,
        'tipo_cargo_choices': Cargo.TIPO_CARGO_CHOICES,
    }

    return render(request, 'planta_docente/gestionar_continuidad.html', contexto)


@login_required
def ver_historial_continuidad_docente(request, docente_pk):
    """Vista para ver el historial completo de continuidad de un docente."""
    docente = get_object_or_404(Docente, pk=docente_pk)

    # Obtener todos los cargos del docente ordenados por fecha
    cargos = Cargo.objects.filter(docente=docente).order_by('fecha_inicio')

    # Construir cadenas de continuidad
    cadenas = []
    cargos_procesados = set()

    for cargo in cargos:
        if cargo.pk in cargos_procesados:
            continue

        # Encontrar el inicio de la cadena
        inicio_cadena = cargo
        while inicio_cadena.cargo_anterior:
            inicio_cadena = inicio_cadena.cargo_anterior

        # Construir la cadena completa desde el inicio
        cadena_actual = []
        cargo_temp = inicio_cadena

        while cargo_temp:
            cadena_actual.append(cargo_temp)
            cargos_procesados.add(cargo_temp.pk)

            # Buscar siguiente
            if hasattr(cargo_temp, 'cargo_siguiente') and cargo_temp.cargo_siguiente:
                cargo_temp = cargo_temp.cargo_siguiente
            else:
                cargo_temp = None

        if cadena_actual:
            cadenas.append(cadena_actual)

    contexto = {
        'docente': docente,
        'cadenas': cadenas,
        'total_cargos': cargos.count(),
    }

    return render(request, 'planta_docente/historial_continuidad.html', contexto)


@login_required
@staff_member_required
def gestionar_mayor_jerarquia_cargo(request, pk):
    """Vista para gestionar vinculación cargo base-temporal M.J."""
    cargo = get_object_or_404(Cargo, pk=pk)

    if request.method == 'POST':
        accion = request.POST.get('accion')

        try:
            if accion == 'vincular':
                # Detectar si es cargo docente o gestión
                cargo_mj_id = request.POST.get('cargo_mj_id')
                tipo_vinculacion = request.POST.get(
                    'tipo_vinculacion')  # 'docente' o 'gestion'

                fecha_inicio_str = request.POST.get('fecha_inicio')
                from datetime import datetime
                fecha_inicio = datetime.strptime(
                    fecha_inicio_str, '%Y-%m-%d').date() if fecha_inicio_str else None

                if tipo_vinculacion == 'docente' and cargo_mj_id:
                    # CASO 1: Vincular con cargo docente
                    cargo_mj = get_object_or_404(Cargo, pk=cargo_mj_id)

                    exito, mensaje = cargo.vincular_cargo_mayor_jerarquia(
                        cargo_mj=cargo_mj,
                        fecha_inicio=fecha_inicio,
                        usuario=request.user
                    )

                elif tipo_vinculacion == 'gestion':
                    # CASO 2: Cargo de gestión
                    tipo_cargo = request.POST.get('tipo_cargo_gestion')
                    descripcion = request.POST.get('descripcion_cargo_gestion')
                    institucion = request.POST.get('institucion_cargo_gestion')

                    if not descripcion:
                        messages.error(
                            request, 'Debe proporcionar descripción del cargo de gestión')
                        return redirect('planta_docente:gestionar_mayor_jerarquia_cargo', pk=pk)

                    exito, mensaje = cargo.vincular_cargo_mayor_jerarquia(
                        fecha_inicio=fecha_inicio,
                        tipo_cargo=tipo_cargo,
                        descripcion_cargo=descripcion,
                        institucion=institucion,
                        usuario=request.user
                    )

                else:
                    messages.error(
                        request, 'Debe seleccionar tipo de vinculación')
                    return redirect('planta_docente:gestionar_mayor_jerarquia_cargo', pk=pk)

                if exito:
                    messages.success(request, mensaje)
                else:
                    messages.error(request, mensaje)

            elif accion == 'desvincular':
                # ... código existente ...
                pass

        except Exception as e:
            messages.error(request, f'Error: {str(e)}')

        return redirect('planta_docente:detalle_cargo', pk=pk)

    # GET request
    info_mj = cargo.get_info_mayor_jerarquia()

    cargos_posibles_mj = Cargo.objects.filter(
        docente=cargo.docente,
        estado='activo'
    ).exclude(pk=cargo.pk).exclude(
        es_cargo_mayor_jerarquia=True
    )

    if cargo.en_licencia_mayor_jerarquia:
        cargos_posibles_mj = cargos_posibles_mj.exclude(
            en_licencia_mayor_jerarquia=True
        ).exclude(
            en_licencia_normal=True
        )

    cargos_posibles_mj = cargos_posibles_mj.order_by('-categoria')

    contexto = {
        'cargo': cargo,
        'info_mj': info_mj,
        'cargos_posibles_mj': cargos_posibles_mj,
        'tipo_cargo_choices': Cargo.TIPO_CARGO_MJ_CHOICES,
    }

    return render(request, 'planta_docente/gestionar_mayor_jerarquia.html', contexto)


@login_required
@staff_member_required
def editar_licencia_mayor_jerarquia_view(request, pk):
    """
    Editar datos de una licencia por mayor jerarquía.
    Permite actualizar: cargo temporal, resolución CSU, fechas.
    """
    cargo_base = get_object_or_404(Cargo, pk=pk)

    # Verificar que tiene licencia M.J.
    if not cargo_base.en_licencia_mayor_jerarquia:
        messages.error(
            request, "Este cargo no tiene una licencia por mayor jerarquía registrada.")
        return redirect("planta_docente:detalle_cargo", pk=pk)

    if request.method == "POST":
        try:
            from datetime import datetime
            from django.core.exceptions import ValidationError

            # Obtener datos del formulario
            resolucion_csu_id = request.POST.get("resolucion_csu")
            fecha_inicio = request.POST.get("fecha_inicio")
            fecha_fin = request.POST.get("fecha_fin")

            # Actualizar resolución CSU (puede ser None/vacío)
            if resolucion_csu_id:
                cargo_base.resolucion_csu = Resolucion.objects.get(
                    pk=resolucion_csu_id)
            else:
                cargo_base.resolucion_csu = None

            # Actualizar fechas de licencia
            if fecha_inicio:
                cargo_base.fecha_inicio_licencia_mj = datetime.strptime(
                    fecha_inicio, '%Y-%m-%d').date()

            if fecha_fin:
                cargo_base.fecha_fin_licencia_mj = datetime.strptime(
                    fecha_fin, '%Y-%m-%d').date()
            elif fecha_fin == '':  # Si viene vacío, limpiar la fecha
                cargo_base.fecha_fin_licencia_mj = None

            cargo_base.save()

            messages.success(
                request, "✓ Licencia por mayor jerarquía actualizada exitosamente.")
            return redirect("planta_docente:detalle_cargo", pk=pk)

        except Resolucion.DoesNotExist:
            messages.error(request, "La resolución seleccionada no existe.")
        except ValidationError as e:
            messages.error(request, str(e))
        except Exception as e:
            messages.error(
                request, f"Error al actualizar la licencia: {str(e)}")

    # GET request - Obtener resoluciones CSU disponibles
    resoluciones_csu = Resolucion.objects.filter(
        origen='csu'
    ).order_by('-año', '-numero')

    from datetime import datetime

    context = {
        "cargo": cargo_base,
        "resoluciones_csu": resoluciones_csu,
        "current_year": datetime.now().year,  # ← NUEVO
    }

    return render(request, "planta_docente/editar_licencia_mj.html", context)


@login_required
@staff_member_required
@require_POST
def crear_resolucion_csu_ajax(request):
    """
    Vista AJAX para crear una nueva resolución CSU.
    Retorna JSON con la información de la resolución creada.
    """
    try:
        # Obtener datos del formulario
        numero = int(request.POST.get('numero'))
        año = int(request.POST.get('año'))
        objeto = request.POST.get('objeto')
        file = request.FILES.get('file')
        cargo_id = request.POST.get('cargo_id')

        # Validar que cargo_id esté presente
        if not cargo_id:
            return JsonResponse({
                'success': False,
                'message': 'Error: El cargo es obligatorio para crear una resolución'
            }, status=400)

        # Obtener el cargo
        try:
            cargo = Cargo.objects.get(pk=cargo_id)
        except Cargo.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': f'Error: Cargo con ID {cargo_id} no encontrado'
            }, status=404)

        # Crear resolución vinculada al cargo
        resolucion = Resolucion.objects.create(
            cargo=cargo,
            numero=numero,
            año=año,
            objeto=objeto,
            origen='csu',
            file=file if file else None
        )

        return JsonResponse({
            'success': True,
            'message': f'Resolución CSU {numero}/{año} creada exitosamente para {cargo.docente.apellido}',
            'resolucion_id': resolucion.pk,
            'numero': resolucion.numero,
            'año': resolucion.año,
            'objeto': resolucion.get_objeto_display() if objeto else '',
        })

    except ValueError as e:
        return JsonResponse({
            'success': False,
            'message': f'Error en los datos: {str(e)}'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error al crear resolución: {str(e)}'
        }, status=500)


@require_GET
@login_required
def asignatura_info_api(request, asignatura_id):
    """
    Endpoint API simple para obtener información de una asignatura.
    Usado para validación dinámica de comisiones en formularios.
    """
    try:
        asignatura = Asignatura.objects.get(pk=asignatura_id)
        return JsonResponse({
            'id': asignatura.pk,
            'nombre': asignatura.nombre,
            'numero_comisiones': asignatura.numero_comisiones or 1,
            'hora_semanal': asignatura.hora_semanal,
            'nivel': asignatura.get_nivel_display(),
            'departamento': asignatura.get_departamento_display(),
        })
    except Asignatura.DoesNotExist:
        return JsonResponse({
            'error': 'Asignatura no encontrada'
        }, status=404)
